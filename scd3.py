import csv
import os
import glob
from datetime import datetime, timedelta
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class WorkScheduleGenerator:
    def __init__(self):
        self.staff_data = {}
        self.schedule = {}
        self.staff_work_hours = defaultdict(float)
        self.daily_lab_assignments = {}  # New: Track daily lab assignments
        
    def parse_staff_availability(self, staff_info):
        """Parse staff availability from CSV file without pandas"""
        csv_file_path = None
        priority_files = ["lich.csv", "schedule.csv", "staff.csv", "data.csv"]
        
        for filename in priority_files:
            if os.path.exists(filename):
                csv_file_path = filename
                break
        
        if csv_file_path is None:
            csv_files = glob.glob("*.csv")
            if csv_files:
                csv_file_path = csv_files[0]
            else:
                raise FileNotFoundError("Không tìm thấy file CSV nào!")
        
        if not os.path.exists(csv_file_path):
            raise FileNotFoundError(f"Không tìm thấy file: {csv_file_path}")
        
        self.staff_data = {}
        
        with open(csv_file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            header = next(reader)
            
            for row in reader:
                if len(row) < 7:
                    continue
                    
                staff_name = row[1].strip()
                if not staff_name:
                    continue
                
                staff_schedule = {'default': {}}
                
                weekday_columns = {
                    2: row[2],  # Thứ 2 (Monday)
                    3: row[3],  # Thứ 3 (Tuesday) 
                    4: row[4],  # Thứ 4 (Wednesday)
                    5: row[5],  # Thứ 5 (Thursday)
                    6: row[6]   # Thứ 6 (Friday)
                }
                
                for weekday_num, shifts_str in weekday_columns.items():
                    if not shifts_str or not shifts_str.strip():
                        continue
                        
                    time_ranges = []
                    shifts = [shift.strip() for shift in str(shifts_str).split(';')]
                    
                    for shift in shifts:
                        if 'Ca 9h - 12h' in shift:
                            time_ranges.append((9, 12))
                        elif 'Ca 13h30 - 16h' in shift:
                            time_ranges.append((13.5, 16))
                        elif 'Ca 16h - 18h30' in shift:
                            time_ranges.append((16, 18.5))
                    
                    if time_ranges:
                        staff_schedule['default'][weekday_num] = time_ranges
                
                if staff_schedule['default']:
                    self.staff_data[staff_name] = staff_schedule
    
    def get_staff_availability(self, staff_name, date, shift_type):
        """Get staff availability for a specific date and shift"""
        staff_info = self.staff_data.get(staff_name, {})
        weekday = date.weekday() + 2
        
        if 'excluded_dates' in staff_info:
            if date.strftime('%Y-%m-%d') in staff_info['excluded_dates']:
                return False, 0
        
        if 'periods' in staff_info:
            for (start_date, end_date), schedule in staff_info['periods'].items():
                if start_date <= date.strftime('%Y-%m-%d') <= end_date:
                    if weekday in schedule:
                        return self._check_shift_overlap(schedule[weekday], shift_type)
            return False, 0
        
        if 'default' in staff_info and weekday in staff_info['default']:
            return self._check_shift_overlap(staff_info['default'][weekday], shift_type)
        
        return False, 0
    
    def _check_shift_overlap(self, time_ranges, shift_type):
        """Check if staff can work in the specified shift and return overlap hours"""
        if shift_type == 'morning':
            shift_start, shift_end = 9, 12
        elif shift_type == 'afternoon1':
            shift_start, shift_end = 13.5, 16
        else:  # afternoon2
            shift_start, shift_end = 16, 18.5
        
        total_overlap = 0
        
        for start, end in time_ranges:
            overlap_start = max(start, shift_start)
            overlap_end = min(end, shift_end)
            if overlap_start < overlap_end:
                total_overlap += overlap_end - overlap_start
        
        return total_overlap > 0, total_overlap
    
    def generate_schedule(self, start_date_str, end_date_str):
        """Generate work schedule with better load balancing and consistent lab assignments"""
        self.parse_staff_availability({})
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        # Reset work hours tracking
        self.staff_work_hours = defaultdict(float)
        
        initial_schedule = self._generate_initial_schedule(start_date, end_date)
        balanced_schedule = self._balance_workload(initial_schedule, start_date, end_date)
        
        return balanced_schedule
    
    def _generate_initial_schedule(self, start_date, end_date):
        """Generate initial schedule with consistent daily lab assignments"""
        current_date = start_date
        schedule = []
        
        while current_date <= end_date:
            if current_date.weekday() < 5:
                date_str = current_date.strftime('%Y-%m-%d')
                
                # Generate staff assignments for the entire day first
                daily_assignments = self._assign_daily_labs(current_date)
                
                # Create schedule entry with consistent lab assignments
                schedule.append({
                    'date': current_date,
                    'Ngày': current_date.strftime('%d/%m/%Y'),
                    'Thứ': f"Thứ {current_date.weekday() + 2}",
                    'assignments': daily_assignments
                })
            
            current_date += timedelta(days=1)
        
        return schedule
    
    def _assign_daily_labs(self, date):
        """Assign labs for the entire day, ensuring consistency across shifts"""
        # Get all available staff for each shift
        morning_available = self._get_available_staff(date, 'morning')
        afternoon1_available = self._get_available_staff(date, 'afternoon1')
        afternoon2_available = self._get_available_staff(date, 'afternoon2')
        
        # Find staff who can work multiple shifts
        all_day_staff = set(morning_available) & set(afternoon1_available) & set(afternoon2_available)
        morning_afternoon1 = set(morning_available) & set(afternoon1_available)
        morning_afternoon2 = set(morning_available) & set(afternoon2_available)
        afternoon1_afternoon2 = set(afternoon1_available) & set(afternoon2_available)
        
        assignments = {
            'morning': {'lab01': '', 'lab02': '', 'lab03': ''},
            'afternoon1': {'lab01': '', 'lab02': '', 'lab03': ''},
            'afternoon2': {'lab01': '', 'lab02': '', 'lab03': ''}
        }
        
        assigned_staff = set()
        
        # Priority 1: Assign staff who can work all day
        available_all_day = list(all_day_staff - assigned_staff)
        for i, staff in enumerate(available_all_day[:3]):
            lab = f'lab0{i+1}'
            assignments['morning'][lab] = staff
            assignments['afternoon1'][lab] = staff
            assignments['afternoon2'][lab] = staff
            assigned_staff.add(staff)
            
            # Update work hours
            morning_can, morning_hours = self.get_staff_availability(staff, date, 'morning')
            afternoon1_can, afternoon1_hours = self.get_staff_availability(staff, date, 'afternoon1')
            afternoon2_can, afternoon2_hours = self.get_staff_availability(staff, date, 'afternoon2')
            
            self.staff_work_hours[staff] += min(morning_hours, 3) + min(afternoon1_hours, 2.5) + min(afternoon2_hours, 2.5)
        
        # Priority 2: Assign staff who can work morning and afternoon1
        remaining_labs = [lab for lab, staff in assignments['morning'].items() if not staff]
        available_morning_afternoon1 = list((morning_afternoon1 - assigned_staff))
        
        for i, staff in enumerate(available_morning_afternoon1):
            if i >= len(remaining_labs):
                break
            lab = remaining_labs[i]
            assignments['morning'][lab] = staff
            assignments['afternoon1'][lab] = staff
            assigned_staff.add(staff)
            
            # Update work hours
            morning_can, morning_hours = self.get_staff_availability(staff, date, 'morning')
            afternoon1_can, afternoon1_hours = self.get_staff_availability(staff, date, 'afternoon1')
            self.staff_work_hours[staff] += min(morning_hours, 3) + min(afternoon1_hours, 2.5)
        
        # Priority 3: Assign staff who can work morning and afternoon2
        remaining_labs = [lab for lab, staff in assignments['morning'].items() if not staff]
        available_morning_afternoon2 = list((morning_afternoon2 - assigned_staff))
        
        for i, staff in enumerate(available_morning_afternoon2):
            if i >= len(remaining_labs):
                break
            lab = remaining_labs[i]
            assignments['morning'][lab] = staff
            assignments['afternoon2'][lab] = staff
            assigned_staff.add(staff)
            
            # Update work hours
            morning_can, morning_hours = self.get_staff_availability(staff, date, 'morning')
            afternoon2_can, afternoon2_hours = self.get_staff_availability(staff, date, 'afternoon2')
            self.staff_work_hours[staff] += min(morning_hours, 3) + min(afternoon2_hours, 2.5)
        
        # Priority 4: Assign staff who can work afternoon1 and afternoon2
        remaining_afternoon_labs = [lab for lab, staff in assignments['afternoon1'].items() if not staff]
        available_afternoon1_afternoon2 = list((afternoon1_afternoon2 - assigned_staff))
        
        for i, staff in enumerate(available_afternoon1_afternoon2):
            if i >= len(remaining_afternoon_labs):
                break
            lab = remaining_afternoon_labs[i]
            assignments['afternoon1'][lab] = staff
            assignments['afternoon2'][lab] = staff
            assigned_staff.add(staff)
            
            # Update work hours
            afternoon1_can, afternoon1_hours = self.get_staff_availability(staff, date, 'afternoon1')
            afternoon2_can, afternoon2_hours = self.get_staff_availability(staff, date, 'afternoon2')
            self.staff_work_hours[staff] += min(afternoon1_hours, 2.5) + min(afternoon2_hours, 2.5)
        
        # Fill remaining slots with available staff
        self._fill_remaining_slots(assignments, date, assigned_staff, morning_available, afternoon1_available, afternoon2_available)
        
        return assignments
    
    def _get_available_staff(self, date, shift_type):
        """Get list of available staff for a specific shift"""
        available = []
        for staff_name in self.staff_data.keys():
            can_work, _ = self.get_staff_availability(staff_name, date, shift_type)
            if can_work:
                available.append(staff_name)
        return available
    
    def _fill_remaining_slots(self, assignments, date, assigned_staff, morning_available, afternoon1_available, afternoon2_available):
        """Fill remaining empty slots"""
        shifts_data = [
            ('morning', morning_available, 3),
            ('afternoon1', afternoon1_available, 2.5),
            ('afternoon2', afternoon2_available, 2.5)
        ]
        
        for shift_type, available_staff, shift_hours in shifts_data:
            remaining_labs = [lab for lab, staff in assignments[shift_type].items() if not staff]
            unassigned_staff = [staff for staff in available_staff if staff not in assigned_staff]
            
            # Sort by workload balance
            unassigned_staff.sort(key=lambda x: self.staff_work_hours[x])
            
            for i, lab in enumerate(remaining_labs):
                if i < len(unassigned_staff):
                    staff = unassigned_staff[i]
                    assignments[shift_type][lab] = staff
                    assigned_staff.add(staff)
                    
                    # Update work hours
                    can_work, hours = self.get_staff_availability(staff, date, shift_type)
                    self.staff_work_hours[staff] += min(hours, shift_hours)
    
    def _balance_workload(self, schedule, start_date, end_date):
        """Balance workload using reassignment while maintaining lab consistency"""
        max_iterations = 5
        improvement_threshold = 0.5  # Minimum improvement to continue balancing
        
        print("Bắt đầu cân bằng workload...")
        
        for iteration in range(max_iterations):
            # Calculate current workload statistics
            current_work_hours = self.staff_work_hours.copy()
            
            if not current_work_hours:
                break
                
            # Calculate workload variance
            work_values = list(current_work_hours.values())
            if len(work_values) <= 1:
                break
                
            avg_hours = sum(work_values) / len(work_values)
            variance_before = sum((hours - avg_hours) ** 2 for hours in work_values) / len(work_values)
            
            # Try to reassign shifts to balance workload
            improved = self._reassign_shifts(schedule, current_work_hours)
            
            if not improved:
                print(f"Không thể cải thiện thêm sau {iteration + 1} vòng lặp")
                break
                
            # Calculate new variance
            new_work_values = list(self.staff_work_hours.values())
            new_avg_hours = sum(new_work_values) / len(new_work_values)
            variance_after = sum((hours - new_avg_hours) ** 2 for hours in new_work_values) / len(new_work_values)
            
            # Check if improvement is significant
            improvement = variance_before - variance_after
            if improvement < improvement_threshold:
                print(f"Cải thiện không đáng kể ({improvement:.2f}), dừng tối ưu hóa")
                break
                
            print(f"Vòng {iteration + 1}: Phương sai giảm từ {variance_before:.2f} xuống {variance_after:.2f}")
        
        # Print final workload distribution
        print("\nPhân bổ giờ làm cuối cùng:")
        for staff, hours in sorted(self.staff_work_hours.items()):
            print(f"  {staff}: {hours:.1f} giờ")
        
        return schedule
    
    def _reassign_shifts(self, schedule, work_hours):
        """Try to reassign shifts to balance workload while maintaining lab consistency"""
        if not work_hours:
            return False
            
        # Find overloaded and underloaded staff
        work_values = list(work_hours.values())
        avg_hours = sum(work_values) / len(work_values)
        
        overloaded_staff = []
        underloaded_staff = []
        
        for staff, hours in work_hours.items():
            if hours > avg_hours + 2:  # Threshold for overloaded
                overloaded_staff.append((staff, hours))
            elif hours < avg_hours - 2:  # Threshold for underloaded
                underloaded_staff.append((staff, hours))
        
        if not overloaded_staff or not underloaded_staff:
            return False
        
        # Sort by workload difference
        overloaded_staff.sort(key=lambda x: x[1], reverse=True)
        underloaded_staff.sort(key=lambda x: x[1])
        
        improvements_made = False
        
        # Try to reassign shifts
        for day_info in schedule:
            date = day_info['date']
            assignments = day_info['assignments']
            
            # Check each shift type
            for shift_type in ['morning', 'afternoon1', 'afternoon2']:
                shift_hours = 3 if shift_type == 'morning' else 2.5
                
                # Look for reassignment opportunities
                for lab in ['lab01', 'lab02', 'lab03']:
                    current_staff = assignments[shift_type][lab]
                    
                    if not current_staff:
                        continue
                        
                    # Check if current staff is overloaded
                    current_staff_hours = work_hours.get(current_staff, 0)
                    if current_staff_hours <= avg_hours + 1:
                        continue
                    
                    # Find a suitable replacement from underloaded staff
                    best_replacement = None
                    best_replacement_hours = float('inf')
                    
                    for replacement_staff, replacement_hours in underloaded_staff:
                        # Check if replacement can work this shift
                        can_work, available_hours = self.get_staff_availability(replacement_staff, date, shift_type)
                        
                        if not can_work or available_hours < shift_hours:
                            continue
                        
                        # Check if this replacement would improve balance
                        if replacement_hours < best_replacement_hours and replacement_hours < current_staff_hours - 1:
                            # Additional check: ensure this doesn't conflict with existing assignments
                            if not self._would_create_conflict(assignments, replacement_staff, shift_type, lab, date):
                                best_replacement = replacement_staff
                                best_replacement_hours = replacement_hours
                    
                    # Make the reassignment if beneficial
                    if best_replacement:
                        # Update assignments
                        assignments[shift_type][lab] = best_replacement
                        
                        # Update work hours tracking
                        work_hours[current_staff] -= shift_hours
                        work_hours[best_replacement] += shift_hours
                        self.staff_work_hours[current_staff] -= shift_hours
                        self.staff_work_hours[best_replacement] += shift_hours
                        
                        # Update the underloaded/overloaded lists
                        underloaded_staff = [(staff, hours) for staff, hours in underloaded_staff 
                                           if staff != best_replacement]
                        if work_hours[best_replacement] < avg_hours - 2:
                            underloaded_staff.append((best_replacement, work_hours[best_replacement]))
                        
                        if work_hours[current_staff] < avg_hours + 2:
                            overloaded_staff = [(staff, hours) for staff, hours in overloaded_staff 
                                              if staff != current_staff]
                        
                        improvements_made = True
                        
                        # Try to maintain lab consistency by reassigning other shifts of the same lab
                        self._try_maintain_lab_consistency(assignments, date, lab, best_replacement, current_staff)
        
        return improvements_made
    
    def _would_create_conflict(self, assignments, staff, shift_type, lab, date):
        """Check if assigning staff to a shift would create conflicts"""
        # Check if staff is already assigned to another lab in the same shift
        for other_lab in ['lab01', 'lab02', 'lab03']:
            if other_lab != lab and assignments[shift_type][other_lab] == staff:
                return True
        
        # Check if staff is already assigned to a different lab in other shifts of the same day
        # This is actually allowed, but we want to maintain some consistency
        other_shifts = [s for s in ['morning', 'afternoon1', 'afternoon2'] if s != shift_type]
        staff_other_labs = set()
        
        for other_shift in other_shifts:
            for other_lab in ['lab01', 'lab02', 'lab03']:
                if assignments[other_shift][other_lab] == staff:
                    staff_other_labs.add(other_lab)
        
        # Prefer to assign to the same lab if staff is already working other shifts
        if staff_other_labs and lab not in staff_other_labs:
            # Check if the preferred lab is available
            for preferred_lab in staff_other_labs:
                if not assignments[shift_type][preferred_lab]:
                    return True  # Prefer the consistent lab assignment
        
        return False
    
    def _try_maintain_lab_consistency(self, assignments, date, lab, new_staff, old_staff):
        """Try to maintain lab consistency by reassigning other shifts"""
        shifts = ['morning', 'afternoon1', 'afternoon2']
        
        for shift_type in shifts:
            if assignments[shift_type][lab] == old_staff:
                # Check if new_staff can work this shift too
                can_work, _ = self.get_staff_availability(new_staff, date, shift_type)
                if can_work:
                    # Check if this wouldn't create conflicts
                    if not self._would_create_conflict(assignments, new_staff, shift_type, lab, date):
                        # Make the reassignment for consistency
                        assignments[shift_type][lab] = new_staff
                        
                        # Update work hours
                        shift_hours = 3 if shift_type == 'morning' else 2.5
                        self.staff_work_hours[old_staff] -= shift_hours
                        self.staff_work_hours[new_staff] += shift_hours
    
    def save_to_excel(self, schedule, filename):
        """Save schedule to Excel file with formatting"""
        if not EXCEL_AVAILABLE:
            # Fallback to CSV if openpyxl not available
            csv_filename = filename.replace('.xlsx', '.csv')
            self.save_to_csv(schedule, csv_filename)
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Schedule"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        headers = ['Date', 'Thứ', 'TYPE', 'Lab01', 'Lab02', 'Lab03']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Data rows
        row_num = 2
        for day_info in schedule:
            date = day_info['date']
            date_str = date.strftime('%d/%m/%Y')
            weekday_str = f"Thứ {date.weekday() + 2}"
            assignments = day_info['assignments']
            
            # Morning shift
            morning_data = [
                date_str,
                weekday_str,
                'Sáng (09:00~12:00)',
                assignments['morning']['lab01'],
                assignments['morning']['lab02'],
                assignments['morning']['lab03']
            ]
            
            for col, value in enumerate(morning_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
            row_num += 1
            
            # Afternoon shift 1
            afternoon1_data = [
                date_str,
                weekday_str,
                'Chiều (13:30 ~ 16:00)',
                assignments['afternoon1']['lab01'],
                assignments['afternoon1']['lab02'],
                assignments['afternoon1']['lab03']
            ]
            
            for col, value in enumerate(afternoon1_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
            row_num += 1
            
            # Afternoon shift 2
            afternoon2_data = [
                date_str,
                weekday_str,
                'Chiều (16:00 ~ 18:30)',
                assignments['afternoon2']['lab01'],
                assignments['afternoon2']['lab02'],
                assignments['afternoon2']['lab03']
            ]
            
            for col, value in enumerate(afternoon2_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
            row_num += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filename)
        print(f"Lịch làm việc đã được lưu vào file: {filename}")

    def save_to_csv(self, schedule, filename):
        """Save schedule to CSV file as fallback"""
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header
            writer.writerow(['Date', 'Thứ', 'TYPE', 'Lab01', 'Lab02', 'Lab03'])
            
            # Data rows
            for day_info in schedule:
                date = day_info['date']
                date_str = date.strftime('%d/%m/%Y')
                weekday_str = f"Thứ {date.weekday() + 2}"
                assignments = day_info['assignments']
                
                # Morning shift
                writer.writerow([
                    date_str, weekday_str, 'Sáng (09:00~12:00)',
                    assignments['morning']['lab01'],
                    assignments['morning']['lab02'],
                    assignments['morning']['lab03']
                ])
                
                # Afternoon shift 1
                writer.writerow([
                    date_str, weekday_str, 'Chiều (13:30 ~ 16:00)',
                    assignments['afternoon1']['lab01'],
                    assignments['afternoon1']['lab02'],
                    assignments['afternoon1']['lab03']
                ])
                
                # Afternoon shift 2
                writer.writerow([
                    date_str, weekday_str, 'Chiều (16:00 ~ 18:30)',
                    assignments['afternoon2']['lab01'],
                    assignments['afternoon2']['lab02'],
                    assignments['afternoon2']['lab03']
                ])
        
        print(f"Lịch làm việc đã được lưu vào file CSV: {filename}")

def get_auto_date_range():
    """Get automatic date range for schedule"""
    today = datetime.now()
    weekday = today.weekday()
    
    if weekday >= 5:
        days_until_next_monday = (7 - weekday) + 0
        start_date = today + timedelta(days=days_until_next_monday)
        end_date = start_date + timedelta(days=4)
    else:
        start_date = today + timedelta(days=1)
        days_until_friday = 4 - weekday
        if days_until_friday <= 0: 
            days_until_friday += 7
        end_date = today + timedelta(days=days_until_friday)
    
    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

def main():
    """Main function"""
    print("=== HỆ THỐNG TẠO LỊCH LÀM VIỆC ===")
    start_date, end_date = get_auto_date_range()
    print(f"Tạo lịch từ {start_date} đến {end_date}")
    
    scheduler = WorkScheduleGenerator()
    schedule = scheduler.generate_schedule(start_date, end_date)
    scheduler.save_to_excel(schedule, 'work_schedule.xlsx')
    
    print("Hoàn thành!")

if __name__ == "__main__":
    main()
